"""
Interactive Web-Based Reporting System for xanadOS Search & Destroy.

This module provides interactive HTML reports with:
- Plotly.js interactive charts
- Drill-down capabilities
- Responsive design
- PDF/Excel export functionality

Phase 2, Task 2.3.1: Interactive Web-Based Reports
"""

import json
import logging
from dataclasses import dataclass, asdict, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
import base64
from io import BytesIO

try:
    import plotly.graph_objects as go
    import plotly.express as px
    from plotly.subplots import make_subplots

    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    logging.warning("Plotly not installed. Install with: pip install plotly>=5.14.0")

try:
    from jinja2 import Environment, FileSystemLoader, select_autoescape

    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False
    logging.warning("Jinja2 not installed. Install with: pip install jinja2>=3.1.0")

try:
    from weasyprint import HTML as WeasyHTML

    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False
    logging.warning(
        "WeasyPrint not installed. For PDF export: pip install weasyprint>=59.0"
    )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning(
        "OpenPyXL not installed. For Excel export: pip install openpyxl>=3.1.0"
    )


logger = logging.getLogger(__name__)


# ============================================================================
# Data Models
# ============================================================================


@dataclass
class ReportData:
    """Container for report data and metadata."""

    report_id: str
    report_type: str  # executive, threat, performance, compliance
    title: str
    generated_at: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    timeframe_start: str | None = None
    timeframe_end: str | None = None
    data: dict[str, Any] = field(default_factory=dict)
    charts: list[dict] = field(default_factory=list)
    summary: dict[str, Any] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class ChartConfig:
    """Configuration for interactive charts."""

    chart_id: str
    chart_type: str  # line, bar, pie, scatter, heatmap
    title: str
    data: dict[str, Any]
    layout: dict[str, Any] = field(default_factory=dict)
    config: dict[str, Any] = field(
        default_factory=lambda: {
            "displayModeBar": True,
            "responsive": True,
            "displaylogo": False,
            "toImageButtonOptions": {
                "format": "png",
                "filename": "chart",
                "height": 800,
                "width": 1200,
            },
        }
    )


@dataclass
class ExportOptions:
    """Options for report export."""

    format: str  # html, pdf, excel
    filename: str
    include_charts: bool = True
    page_size: str = "A4"  # For PDF: A4, Letter, Legal
    orientation: str = "portrait"  # portrait, landscape
    quality: str = "high"  # low, medium, high


# ============================================================================
# Web Report Generator
# ============================================================================


class WebReportGenerator:
    """
    Generates interactive web-based reports with Plotly charts.

    Features:
    - HTML reports with interactive charts
    - Drill-down capabilities
    - Responsive design (mobile-friendly)
    - PDF/Excel export
    """

    def __init__(self, template_dir: Path | None = None):
        """
        Initialize the web report generator.

        Args:
            template_dir: Directory containing Jinja2 templates
        """
        self.template_dir = template_dir or Path(__file__).parent / "templates"
        self.template_dir.mkdir(parents=True, exist_ok=True)

        if JINJA2_AVAILABLE:
            self.env = Environment(
                loader=FileSystemLoader(str(self.template_dir)),
                autoescape=select_autoescape(["html", "xml"]),
            )
        else:
            self.env = None
            logger.warning("Jinja2 not available, template rendering disabled")

        logger.info(
            f"WebReportGenerator initialized with template dir: {self.template_dir}"
        )

    # ========================================
    # Report Generation
    # ========================================

    def generate_executive_report(
        self, scan_results: list[dict], timeframe_days: int = 30
    ) -> ReportData:
        """
        Generate executive dashboard report.

        Args:
            scan_results: List of scan result dictionaries
            timeframe_days: Number of days to include

        Returns:
            ReportData with executive summary
        """
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=timeframe_days)

        # Calculate summary statistics
        total_scans = len(scan_results)
        total_threats = sum(1 for r in scan_results if r.get("threats_found", 0) > 0)
        total_files_scanned = sum(r.get("files_scanned", 0) for r in scan_results)
        total_threats_found = sum(r.get("threats_found", 0) for r in scan_results)

        # Calculate threat severity breakdown
        severity_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for result in scan_results:
            for threat in result.get("threats", []):
                severity = threat.get("severity", "low").lower()
                if severity in severity_counts:
                    severity_counts[severity] += 1

        # Create charts
        charts = []

        # Chart 1: Threat trend over time
        threat_trend_chart = self._create_threat_trend_chart(scan_results)
        charts.append(threat_trend_chart)

        # Chart 2: Severity distribution
        severity_chart = self._create_severity_pie_chart(severity_counts)
        charts.append(severity_chart)

        # Chart 3: Scan performance
        performance_chart = self._create_performance_chart(scan_results)
        charts.append(performance_chart)

        # Build summary
        summary = {
            "total_scans": total_scans,
            "scans_with_threats": total_threats,
            "total_files_scanned": total_files_scanned,
            "total_threats_found": total_threats_found,
            "threat_rate": (
                (total_threats / total_scans * 100) if total_scans > 0 else 0
            ),
            "severity_breakdown": severity_counts,
            "avg_scan_time": self._calculate_avg_scan_time(scan_results),
            "top_threat_types": self._get_top_threat_types(scan_results, limit=5),
        }

        return ReportData(
            report_id=f"exec_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            report_type="executive",
            title="Executive Security Dashboard",
            timeframe_start=start_date.isoformat(),
            timeframe_end=end_date.isoformat(),
            data={"scan_results": scan_results},
            charts=charts,
            summary=summary,
        )

    def generate_threat_analysis_report(
        self, threats: list[dict], timeframe_days: int = 30
    ) -> ReportData:
        """
        Generate detailed threat analysis report.

        Args:
            threats: List of detected threats
            timeframe_days: Number of days to include

        Returns:
            ReportData with threat analysis
        """
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=timeframe_days)

        # Analyze threats
        threat_types = {}
        for threat in threats:
            threat_type = threat.get("type", "Unknown")
            threat_types[threat_type] = threat_types.get(threat_type, 0) + 1

        # Create charts
        charts = []

        if PLOTLY_AVAILABLE:
            # Chart 1: Threat types distribution
            type_chart = self._create_threat_type_chart(threat_types)
            charts.append(type_chart)

            # Chart 2: Threat timeline
            timeline_chart = self._create_threat_timeline_chart(threats)
            charts.append(timeline_chart)

            # Chart 3: Threat sources (file paths)
            sources_chart = self._create_threat_sources_chart(threats)
            charts.append(sources_chart)

        summary = {
            "total_threats": len(threats),
            "unique_threat_types": len(threat_types),
            "most_common_threat": (
                max(threat_types.items(), key=lambda x: x[1])[0]
                if threat_types
                else "None"
            ),
            "threat_type_breakdown": threat_types,
            "avg_threats_per_day": (
                len(threats) / timeframe_days if timeframe_days > 0 else 0
            ),
        }

        return ReportData(
            report_id=f"threat_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            report_type="threat",
            title="Threat Analysis Report",
            timeframe_start=start_date.isoformat(),
            timeframe_end=end_date.isoformat(),
            data={"threats": threats},
            charts=charts,
            summary=summary,
        )

    def generate_performance_report(
        self, performance_data: list[dict], timeframe_days: int = 30
    ) -> ReportData:
        """
        Generate system performance report.

        Args:
            performance_data: List of performance metrics
            timeframe_days: Number of days to include

        Returns:
            ReportData with performance analysis
        """
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=timeframe_days)

        # Calculate performance metrics
        avg_cpu = (
            sum(d.get("cpu_percent", 0) for d in performance_data)
            / len(performance_data)
            if performance_data
            else 0
        )
        avg_memory = (
            sum(d.get("memory_mb", 0) for d in performance_data) / len(performance_data)
            if performance_data
            else 0
        )
        avg_scan_time = (
            sum(d.get("scan_duration", 0) for d in performance_data)
            / len(performance_data)
            if performance_data
            else 0
        )

        # Create charts
        charts = []

        # Chart 1: Resource usage over time
        resource_chart = self._create_resource_usage_chart(performance_data)
        charts.append(resource_chart)

        # Chart 2: Scan throughput
        throughput_chart = self._create_throughput_chart(performance_data)
        charts.append(throughput_chart)

        # Chart 3: Cache efficiency
        cache_chart = self._create_cache_efficiency_chart(performance_data)
        charts.append(cache_chart)

        summary = {
            "avg_cpu_percent": round(avg_cpu, 2),
            "avg_memory_mb": round(avg_memory, 2),
            "avg_scan_duration_seconds": round(avg_scan_time, 2),
            "total_data_points": len(performance_data),
            "peak_cpu": max(
                (d.get("cpu_percent", 0) for d in performance_data), default=0
            ),
            "peak_memory_mb": max(
                (d.get("memory_mb", 0) for d in performance_data), default=0
            ),
        }

        return ReportData(
            report_id=f"perf_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            report_type="performance",
            title="Performance Analysis Report",
            timeframe_start=start_date.isoformat(),
            timeframe_end=end_date.isoformat(),
            data={"performance_data": performance_data},
            charts=charts,
            summary=summary,
        )

    def generate_compliance_report(
        self, framework: str, compliance_data: dict
    ) -> ReportData:
        """
        Generate compliance audit report.

        Args:
            framework: Compliance framework name (PCI_DSS, NIST_CSF, etc.)
            compliance_data: Compliance assessment data

        Returns:
            ReportData with compliance status
        """
        # Calculate compliance score
        total_controls = compliance_data.get("total_controls", 0)
        passed_controls = compliance_data.get("passed_controls", 0)
        compliance_score = (
            (passed_controls / total_controls * 100) if total_controls > 0 else 0
        )

        # Create charts
        charts = []

        if PLOTLY_AVAILABLE:
            # Chart 1: Compliance score gauge
            gauge_chart = self._create_compliance_gauge(compliance_score)
            charts.append(gauge_chart)

            # Chart 2: Control category breakdown
            category_chart = self._create_control_category_chart(compliance_data)
            charts.append(category_chart)

            # Chart 3: Gaps and remediation priorities
            gap_chart = self._create_gap_analysis_chart(compliance_data)
            charts.append(gap_chart)

        summary = {
            "framework": framework,
            "compliance_score": round(compliance_score, 2),
            "total_controls": total_controls,
            "passed_controls": passed_controls,
            "failed_controls": total_controls - passed_controls,
            "critical_gaps": len(compliance_data.get("critical_gaps", [])),
            "recommendations": compliance_data.get("recommendations", []),
        }

        return ReportData(
            report_id=f"comp_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            report_type="compliance",
            title=f"{framework} Compliance Audit",
            data={"compliance_data": compliance_data},
            charts=charts,
            summary=summary,
        )

    # ========================================
    # Chart Creation (Plotly)
    # ========================================

    def _create_threat_trend_chart(self, scan_results: list[dict]) -> dict:
        """Create line chart showing threat detection trend."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "threat_trend", "html": "<p>Plotly not available</p>"}

        # Group by date
        daily_threats = {}
        for result in scan_results:
            date_str = result.get("timestamp", "")[:10]  # YYYY-MM-DD
            threats = result.get("threats_found", 0)
            daily_threats[date_str] = daily_threats.get(date_str, 0) + threats

        dates = sorted(daily_threats.keys())
        counts = [daily_threats[d] for d in dates]

        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=dates,
                y=counts,
                mode="lines+markers",
                name="Threats Detected",
                line=dict(color="#e74c3c", width=2),
                marker=dict(size=8),
            )
        )

        fig.update_layout(
            title="Threat Detection Trend",
            xaxis_title="Date",
            yaxis_title="Threats Detected",
            hovermode="x unified",
            template="plotly_white",
        )

        return {
            "chart_id": "threat_trend",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="threat_trend"),
        }

    def _create_severity_pie_chart(self, severity_counts: dict) -> dict:
        """Create pie chart for severity distribution."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "severity_pie", "html": "<p>Plotly not available</p>"}

        colors = {
            "critical": "#8b0000",
            "high": "#e74c3c",
            "medium": "#f39c12",
            "low": "#f1c40f",
        }

        labels = list(severity_counts.keys())
        values = list(severity_counts.values())
        chart_colors = [colors.get(label, "#95a5a6") for label in labels]

        fig = go.Figure(
            data=[
                go.Pie(
                    labels=[l.capitalize() for l in labels],
                    values=values,
                    marker=dict(colors=chart_colors),
                    hole=0.3,
                    textinfo="label+percent",
                )
            ]
        )

        fig.update_layout(title="Threat Severity Distribution", template="plotly_white")

        return {
            "chart_id": "severity_pie",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="severity_pie"),
        }

    def _create_performance_chart(self, scan_results: list[dict]) -> dict:
        """Create chart showing scan performance metrics."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "performance", "html": "<p>Plotly not available</p>"}

        timestamps = [r.get("timestamp", "") for r in scan_results]
        scan_times = [r.get("scan_duration", 0) for r in scan_results]
        files_scanned = [r.get("files_scanned", 0) for r in scan_results]

        fig = make_subplots(
            rows=2,
            cols=1,
            subplot_titles=("Scan Duration", "Files Scanned"),
            vertical_spacing=0.15,
        )

        fig.add_trace(
            go.Scatter(x=timestamps, y=scan_times, mode="lines", name="Scan Time (s)"),
            row=1,
            col=1,
        )

        fig.add_trace(
            go.Bar(x=timestamps, y=files_scanned, name="Files Scanned"), row=2, col=1
        )

        fig.update_layout(
            height=600,
            title_text="Scan Performance Metrics",
            template="plotly_white",
            showlegend=True,
        )

        return {
            "chart_id": "performance",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="performance"),
        }

    def _create_threat_type_chart(self, threat_types: dict) -> dict:
        """Create bar chart for threat types."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "threat_types", "html": "<p>Plotly not available</p>"}

        types = list(threat_types.keys())
        counts = list(threat_types.values())

        fig = go.Figure(data=[go.Bar(x=types, y=counts, marker_color="#3498db")])

        fig.update_layout(
            title="Threat Types Distribution",
            xaxis_title="Threat Type",
            yaxis_title="Count",
            template="plotly_white",
        )

        return {
            "chart_id": "threat_types",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="threat_types"),
        }

    def _create_threat_timeline_chart(self, threats: list[dict]) -> dict:
        """Create timeline chart for threat detection."""
        if not PLOTLY_AVAILABLE:
            return {
                "chart_id": "threat_timeline",
                "html": "<p>Plotly not available</p>",
            }

        # Group by hour
        hourly_counts = {}
        for threat in threats:
            timestamp = threat.get("detected_at", "")
            if timestamp:
                hour = timestamp[:13]  # YYYY-MM-DD HH
                hourly_counts[hour] = hourly_counts.get(hour, 0) + 1

        hours = sorted(hourly_counts.keys())
        counts = [hourly_counts[h] for h in hours]

        fig = go.Figure(
            data=[
                go.Scatter(
                    x=hours,
                    y=counts,
                    mode="lines+markers",
                    fill="tozeroy",
                    line=dict(color="#e74c3c"),
                )
            ]
        )

        fig.update_layout(
            title="Threat Detection Timeline",
            xaxis_title="Time",
            yaxis_title="Threats Detected",
            template="plotly_white",
        )

        return {
            "chart_id": "threat_timeline",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="threat_timeline"),
        }

    def _create_threat_sources_chart(self, threats: list[dict]) -> dict:
        """Create chart showing top threat sources."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "threat_sources", "html": "<p>Plotly not available</p>"}

        # Count threats by directory
        dir_counts = {}
        for threat in threats:
            file_path = threat.get("file_path", "")
            directory = "/".join(file_path.split("/")[:-1]) or "/"
            dir_counts[directory] = dir_counts.get(directory, 0) + 1

        # Top 10 directories
        top_dirs = sorted(dir_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        dirs = [d[0] for d in top_dirs]
        counts = [d[1] for d in top_dirs]

        fig = go.Figure(
            data=[go.Bar(x=counts, y=dirs, orientation="h", marker_color="#9b59b6")]
        )

        fig.update_layout(
            title="Top Threat Sources (Directories)",
            xaxis_title="Threat Count",
            yaxis_title="Directory",
            template="plotly_white",
            height=400,
        )

        return {
            "chart_id": "threat_sources",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="threat_sources"),
        }

    def _create_resource_usage_chart(self, performance_data: list[dict]) -> dict:
        """Create chart for resource usage over time."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "resource_usage", "html": "<p>Plotly not available</p>"}

        timestamps = [d.get("timestamp", "") for d in performance_data]
        cpu = [d.get("cpu_percent", 0) for d in performance_data]
        memory = [
            d.get("memory_mb", 0) / 1024 for d in performance_data
        ]  # Convert to GB

        fig = make_subplots(
            rows=2,
            cols=1,
            subplot_titles=("CPU Usage (%)", "Memory Usage (GB)"),
            vertical_spacing=0.15,
        )

        fig.add_trace(
            go.Scatter(
                x=timestamps,
                y=cpu,
                mode="lines",
                name="CPU",
                line=dict(color="#3498db"),
            ),
            row=1,
            col=1,
        )

        fig.add_trace(
            go.Scatter(
                x=timestamps,
                y=memory,
                mode="lines",
                name="Memory",
                line=dict(color="#e74c3c"),
            ),
            row=2,
            col=1,
        )

        fig.update_layout(
            height=600, title_text="Resource Usage Over Time", template="plotly_white"
        )

        return {
            "chart_id": "resource_usage",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="resource_usage"),
        }

    def _create_throughput_chart(self, performance_data: list[dict]) -> dict:
        """Create chart for scan throughput."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "throughput", "html": "<p>Plotly not available</p>"}

        timestamps = [d.get("timestamp", "") for d in performance_data]
        throughput = [
            (
                d.get("files_scanned", 0) / d.get("scan_duration", 1)
                if d.get("scan_duration", 0) > 0
                else 0
            )
            for d in performance_data
        ]

        fig = go.Figure(
            data=[
                go.Scatter(
                    x=timestamps,
                    y=throughput,
                    mode="lines+markers",
                    name="Files/Second",
                    line=dict(color="#27ae60"),
                )
            ]
        )

        fig.update_layout(
            title="Scan Throughput (Files/Second)",
            xaxis_title="Time",
            yaxis_title="Files per Second",
            template="plotly_white",
        )

        return {
            "chart_id": "throughput",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="throughput"),
        }

    def _create_cache_efficiency_chart(self, performance_data: list[dict]) -> dict:
        """Create chart for cache hit rate."""
        if not PLOTLY_AVAILABLE:
            return {
                "chart_id": "cache_efficiency",
                "html": "<p>Plotly not available</p>",
            }

        timestamps = [d.get("timestamp", "") for d in performance_data]
        hit_rate = [d.get("cache_hit_rate", 0) * 100 for d in performance_data]

        fig = go.Figure(
            data=[
                go.Scatter(
                    x=timestamps,
                    y=hit_rate,
                    mode="lines",
                    fill="tozeroy",
                    name="Cache Hit Rate",
                    line=dict(color="#f39c12"),
                )
            ]
        )

        fig.update_layout(
            title="Cache Efficiency (Hit Rate %)",
            xaxis_title="Time",
            yaxis_title="Hit Rate (%)",
            template="plotly_white",
        )

        return {
            "chart_id": "cache_efficiency",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="cache_efficiency"),
        }

    def _create_compliance_gauge(self, compliance_score: float) -> dict:
        """Create gauge chart for compliance score."""
        if not PLOTLY_AVAILABLE:
            return {
                "chart_id": "compliance_gauge",
                "html": "<p>Plotly not available</p>",
            }

        fig = go.Figure(
            go.Indicator(
                mode="gauge+number+delta",
                value=compliance_score,
                domain={"x": [0, 1], "y": [0, 1]},
                title={"text": "Compliance Score (%)"},
                delta={"reference": 100},
                gauge={
                    "axis": {"range": [None, 100]},
                    "bar": {"color": "darkblue"},
                    "steps": [
                        {"range": [0, 50], "color": "#e74c3c"},
                        {"range": [50, 75], "color": "#f39c12"},
                        {"range": [75, 100], "color": "#27ae60"},
                    ],
                    "threshold": {
                        "line": {"color": "red", "width": 4},
                        "thickness": 0.75,
                        "value": 80,
                    },
                },
            )
        )

        fig.update_layout(height=400, template="plotly_white")

        return {
            "chart_id": "compliance_gauge",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="compliance_gauge"),
        }

    def _create_control_category_chart(self, compliance_data: dict) -> dict:
        """Create chart for control category status."""
        if not PLOTLY_AVAILABLE:
            return {
                "chart_id": "control_categories",
                "html": "<p>Plotly not available</p>",
            }

        categories = compliance_data.get("categories", {})

        category_names = list(categories.keys())
        passed = [cat.get("passed", 0) for cat in categories.values()]
        failed = [cat.get("failed", 0) for cat in categories.values()]

        fig = go.Figure(
            data=[
                go.Bar(
                    name="Passed", x=category_names, y=passed, marker_color="#27ae60"
                ),
                go.Bar(
                    name="Failed", x=category_names, y=failed, marker_color="#e74c3c"
                ),
            ]
        )

        fig.update_layout(
            title="Control Category Status",
            xaxis_title="Category",
            yaxis_title="Controls",
            barmode="stack",
            template="plotly_white",
        )

        return {
            "chart_id": "control_categories",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="control_categories"),
        }

    def _create_gap_analysis_chart(self, compliance_data: dict) -> dict:
        """Create chart showing compliance gaps."""
        if not PLOTLY_AVAILABLE:
            return {"chart_id": "gap_analysis", "html": "<p>Plotly not available</p>"}

        gaps = compliance_data.get("gaps", [])

        # Group by priority
        priority_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        for gap in gaps:
            priority = gap.get("priority", "low").lower()
            if priority in priority_counts:
                priority_counts[priority] += 1

        priorities = list(priority_counts.keys())
        counts = list(priority_counts.values())

        colors = ["#8b0000", "#e74c3c", "#f39c12", "#f1c40f"]

        fig = go.Figure(
            data=[
                go.Bar(
                    x=[p.capitalize() for p in priorities],
                    y=counts,
                    marker_color=colors,
                )
            ]
        )

        fig.update_layout(
            title="Compliance Gaps by Priority",
            xaxis_title="Priority",
            yaxis_title="Number of Gaps",
            template="plotly_white",
        )

        return {
            "chart_id": "gap_analysis",
            "html": fig.to_html(include_plotlyjs="cdn", div_id="gap_analysis"),
        }

    # ========================================
    # HTML Rendering
    # ========================================

    def render_html(self, report_data: ReportData) -> str:
        """
        Render report data to HTML using Jinja2 templates.

        Args:
            report_data: Report data to render

        Returns:
            HTML string
        """
        if not self.env:
            # Fallback: simple HTML without templates
            return self._render_html_fallback(report_data)

        try:
            template_name = f"{report_data.report_type}_report.html"
            template = self.env.get_template(template_name)

            return template.render(
                report=report_data,
                charts=[chart.get("html", "") for chart in report_data.charts],
                generated_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            )
        except Exception as e:
            logger.warning(f"Template rendering failed: {e}, using fallback")
            return self._render_html_fallback(report_data)

    def _render_html_fallback(self, report_data: ReportData) -> str:
        """Fallback HTML rendering without templates."""
        charts_html = "\n".join(chart.get("html", "") for chart in report_data.charts)

        summary_items = "\n".join(
            f"<li><strong>{key}:</strong> {value}</li>"
            for key, value in report_data.summary.items()
        )

        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_data.title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }}
        .summary {{
            background: #ecf0f1;
            padding: 20px;
            border-radius: 6px;
            margin: 20px 0;
        }}
        .summary ul {{
            list-style: none;
            padding: 0;
        }}
        .summary li {{
            padding: 8px 0;
            border-bottom: 1px solid #bdc3c7;
        }}
        .charts {{
            margin-top: 30px;
        }}
        @media (max-width: 768px) {{
            body {{
                padding: 10px;
            }}
            .container {{
                padding: 15px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{report_data.title}</h1>
        <p><em>Generated: {report_data.generated_at}</em></p>

        <div class="summary">
            <h2>Summary</h2>
            <ul>
                {summary_items}
            </ul>
        </div>

        <div class="charts">
            {charts_html}
        </div>
    </div>
</body>
</html>
"""
        return html

    # ========================================
    # Export Functions
    # ========================================

    def export_to_pdf(
        self,
        report_data: ReportData,
        output_path: Path,
        options: ExportOptions | None = None,
    ) -> bool:
        """
        Export report to PDF.

        Args:
            report_data: Report data to export
            output_path: Output PDF file path
            options: Export options

        Returns:
            True if successful
        """
        if not WEASYPRINT_AVAILABLE:
            logger.error("WeasyPrint not installed, cannot export to PDF")
            return False

        try:
            html_content = self.render_html(report_data)

            # Convert HTML to PDF
            WeasyHTML(string=html_content).write_pdf(str(output_path))

            logger.info(f"PDF exported to {output_path}")
            return True
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return False

    def export_to_excel(
        self,
        report_data: ReportData,
        output_path: Path,
        options: ExportOptions | None = None,
    ) -> bool:
        """
        Export report to Excel.

        Args:
            report_data: Report data to export
            output_path: Output Excel file path
            options: Export options

        Returns:
            True if successful
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("OpenPyXL not installed, cannot export to Excel")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            # Add title
            ws["A1"] = report_data.title
            ws["A1"].font = Font(size=16, bold=True)

            # Add summary data
            row = 3
            for key, value in report_data.summary.items():
                ws[f"A{row}"] = key
                ws[f"B{row}"] = str(value)
                row += 1

            # Save workbook
            wb.save(str(output_path))

            logger.info(f"Excel exported to {output_path}")
            return True
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return False

    # ========================================
    # Utility Methods
    # ========================================

    def _calculate_avg_scan_time(self, scan_results: list[dict]) -> float:
        """Calculate average scan duration."""
        if not scan_results:
            return 0.0

        total_time = sum(r.get("scan_duration", 0) for r in scan_results)
        return round(total_time / len(scan_results), 2)

    def _get_top_threat_types(
        self, scan_results: list[dict], limit: int = 5
    ) -> list[dict]:
        """Get top N most common threat types."""
        threat_counts = {}

        for result in scan_results:
            for threat in result.get("threats", []):
                threat_type = threat.get("type", "Unknown")
                threat_counts[threat_type] = threat_counts.get(threat_type, 0) + 1

        sorted_threats = sorted(threat_counts.items(), key=lambda x: x[1], reverse=True)

        return [
            {"type": threat_type, "count": count}
            for threat_type, count in sorted_threats[:limit]
        ]


# ============================================================================
# Main Entry Point (for testing)
# ============================================================================

if __name__ == "__main__":
    # Example usage
    generator = WebReportGenerator()

    # Sample data
    sample_scans = [
        {
            "timestamp": "2025-12-01T10:00:00",
            "scan_duration": 45.2,
            "files_scanned": 1200,
            "threats_found": 3,
            "threats": [
                {
                    "type": "Malware",
                    "severity": "high",
                    "file_path": "/tmp/malware.exe",
                },
                {
                    "type": "Trojan",
                    "severity": "critical",
                    "file_path": "/tmp/trojan.bin",
                },
                {"type": "Adware", "severity": "low", "file_path": "/tmp/adware.dll"},
            ],
        },
        {
            "timestamp": "2025-12-02T10:00:00",
            "scan_duration": 38.5,
            "files_scanned": 1150,
            "threats_found": 1,
            "threats": [
                {
                    "type": "Malware",
                    "severity": "medium",
                    "file_path": "/tmp/suspicious.exe",
                },
            ],
        },
    ]

    # Generate executive report
    exec_report = generator.generate_executive_report(sample_scans, timeframe_days=30)
    html = generator.render_html(exec_report)

    # Save HTML
    output_file = Path("executive_report.html")
    output_file.write_text(html)
    print(f"Report generated: {output_file}")
